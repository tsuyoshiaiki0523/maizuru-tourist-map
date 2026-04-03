#!/usr/bin/env python3
"""
import_timetable.py
bus_timetable.xlsx → index.html の BUS_STOP_TIMETABLE を更新

使い方: python3 import_timetable.py
  - bus_timetable.xlsx を読み込む
  - index.html 内の BUS_STOP_TIMETABLE を差し替え
  - sw.js のキャッシュバージョンを +1
  - バックアップを index.html.bak に保存
"""

import re
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(SCRIPT_DIR, 'index.html')
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'bus_timetable.xlsx')
BACKUP_PATH = os.path.join(SCRIPT_DIR, 'index.html.bak')


def read_excel(path):
    """Excel → { stop_name: { routes: { dir_key: { dest, times } } } }"""
    wb = load_workbook(path, read_only=True, data_only=True)
    timetable = {}

    for ws in wb.worksheets:
        if ws.title == '全停留所一覧':
            continue

        # E1からキーを取得
        key_cell = str(ws['E1'].value or '')
        m = re.search(r'キー:\s*(\S+)', key_cell)
        if not m:
            print(f"  ⚠ シート '{ws.title}': キー情報なし。スキップ。")
            continue
        dir_key = m.group(1)

        for row in ws.iter_rows(min_row=4, max_col=4, values_only=True):
            _, name_val, dest_val, times_val = row
            stop_name = str(name_val or '').strip()
            if not stop_name or stop_name.startswith('■'):
                continue  # 凡例行をスキップ

            times_str = str(times_val or '').strip()
            if not times_str:
                continue

            dest = str(dest_val or '').strip()
            times = []
            for t in times_str.split(','):
                t = t.strip()
                if re.match(r'^\d{1,2}:\d{2}$', t):
                    h, m = t.split(':')
                    times.append(f"{int(h):02d}:{m}")
                elif t:
                    print(f"  ⚠ 無効: '{t}' ({stop_name}, {dir_key})")

            if not times:
                continue

            timetable.setdefault(stop_name, {'routes': {}})
            timetable[stop_name]['routes'][dir_key] = {'dest': dest, 'times': times}

    wb.close()
    return timetable


def generate_js(timetable):
    """Python dict → JavaScript const BUS_STOP_TIMETABLE = { ... };"""
    lines = ['const BUS_STOP_TIMETABLE = {']
    for sn in sorted(timetable.keys()):
        lines.append(f'  "{sn}": {{')
        lines.append('    routes: {')
        for dk in sorted(timetable[sn]['routes'].keys()):
            e = timetable[sn]['routes'][dk]
            ts = ','.join(f'"{t}"' for t in e['times'])
            lines.append(f'      "{dk}": {{ dest: "{e["dest"]}", times: [{ts}] }},')
        lines.append('    }')
        lines.append('  },')
    lines.append('};')
    return '\n'.join(lines)


def update_html(html, new_js):
    marker = 'const BUS_STOP_TIMETABLE'
    si = html.find(marker)
    if si == -1:
        raise ValueError("BUS_STOP_TIMETABLE not found")
    bi = html.index('{', si)
    depth, i = 0, bi
    while i < len(html):
        if html[i] == '{': depth += 1
        elif html[i] == '}':
            depth -= 1
            if depth == 0: break
        elif html[i] in ('"', "'"):
            q = html[i]; i += 1
            while i < len(html) and html[i] != q:
                if html[i] == '\\': i += 1
                i += 1
        i += 1
    ei = i + 1
    if ei < len(html) and html[ei] == ';':
        ei += 1
    return html[:si] + new_js + html[ei:]


def bump_sw(script_dir):
    p = os.path.join(script_dir, 'sw.js')
    if not os.path.exists(p):
        return
    with open(p, 'r') as f:
        s = f.read()
    m = re.search(r"'maizuru-tour-v(\d+)'", s)
    if m:
        ov = int(m.group(1))
        nv = ov + 1
        s = s.replace(f"'maizuru-tour-v{ov}'", f"'maizuru-tour-v{nv}'")
        with open(p, 'w') as f:
            f.write(s)
        print(f"✅ sw.js: v{ov} → v{nv}")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ {EXCEL_PATH} が見つかりません。先に export_timetable.py を実行してください。")
        return

    print(f"読み込み: {EXCEL_PATH}")
    tt = read_excel(EXCEL_PATH)
    sc = len(tt)
    ec = sum(len(s['routes']) for s in tt.values())
    print(f"  停留所: {sc}, エントリ: {ec}")

    if sc == 0:
        print("❌ データ空。Excel確認してください。")
        return

    print("JS生成 → HTML更新...")
    with open(HTML_PATH, 'r') as f:
        html = f.read()

    # 更新前のエントリ数を確認
    old_count = html.count('"times":') + html.count('"times" :')
    # もっと正確にカウント
    import json
    old_marker = 'const BUS_STOP_TIMETABLE'
    old_si = html.find(old_marker)
    if old_si != -1:
        old_bi = html.index('{', old_si)
        depth, i = 0, old_bi
        while i < len(html):
            if html[i] == '{': depth += 1
            elif html[i] == '}':
                depth -= 1
                if depth == 0: break
            elif html[i] in ('"', "'"):
                q = html[i]; i += 1
                while i < len(html) and html[i] != q:
                    if html[i] == '\\': i += 1
                    i += 1
            i += 1

    with open(BACKUP_PATH, 'w') as f:
        f.write(html)
    print(f"  バックアップ: {BACKUP_PATH}")

    new_js = generate_js(tt)
    html = update_html(html, new_js)
    with open(HTML_PATH, 'w') as f:
        f.write(html)

    bump_sw(SCRIPT_DIR)
    print(f"✅ 完了! 停留所: {sc}, エントリ: {ec}")
    print("\n次のステップ:")
    print("  1. ブラウザで確認")
    print("  2. git add index.html sw.js && git commit && git push")


if __name__ == '__main__':
    main()
