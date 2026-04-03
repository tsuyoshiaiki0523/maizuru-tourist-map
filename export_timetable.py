#!/usr/bin/env python3
"""
export_timetable.py
index.html の BUS_LINE_DATA / BUS_STOP_TIMETABLE → bus_timetable.xlsx

構造:
  - 「全停留所一覧」シート: 路線別の停留所・座標一覧
  - 路線×方向シート: 行=停留所（その方向のデータを持つ全停留所）、列=No./停留所名/行先/発車時刻

使い方: python3 export_timetable.py
"""

import re
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(SCRIPT_DIR, 'index.html')
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'bus_timetable.xlsx')


def hex_to_rgb(h):
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def lighten(hex_color, factor=0.7):
    r, g, b = hex_to_rgb(hex_color)
    return f"{int(r+(255-r)*factor):02X}{int(g+(255-g)*factor):02X}{int(b+(255-b)*factor):02X}"


def extract_js_object(html, var_name):
    pattern = rf'const\s+{var_name}\s*=\s*\{{'
    match = re.search(pattern, html)
    if not match:
        raise ValueError(f"{var_name} not found")
    start = match.start() + match.group().index('{')
    depth, i = 0, start
    while i < len(html):
        if html[i] == '{': depth += 1
        elif html[i] == '}':
            depth -= 1
            if depth == 0: break
        elif html[i] in ('"', "'"):
            q = html[i]; i += 1
            while i < len(html) and html[i] != q:
                if html[i] == '\\': i += 1
                i += 1
        i += 1
    js = html[start:i+1]
    # Convert JS to JSON
    result = []
    j = 0
    while j < len(js):
        if js[j] == "'":
            j += 1; s = ''
            while j < len(js) and js[j] != "'":
                if js[j] == '\\': s += js[j:j+2]; j += 2
                else:
                    s += ('\\"' if js[j] == '"' else js[j]); j += 1
            j += 1; result.append(f'"{s}"')
        elif js[j] == '"':
            result.append('"'); j += 1
            while j < len(js) and js[j] != '"':
                if js[j] == '\\': result.append(js[j:j+2]); j += 2
                else: result.append(js[j]); j += 1
            if j < len(js): result.append('"'); j += 1
        else: result.append(js[j]); j += 1
    text = ''.join(result)
    text = re.sub(r'(?<=[{,\n])\s*(\w[\w\-]*)\s*:', r' "\1":', text)
    text = re.sub(r',\s*([}\]])', r'\1', text)
    return json.loads(text)


def get_direction_info(route_key, suffix):
    d = {
        'eastWest': {'_L': '左回り(西舞鶴→東舞鶴)', '_R': '右回り(東舞鶴→西舞鶴)'},
        'mihama': {'_out': '往路(東舞鶴→平)', '_in': '復路(平→東舞鶴)', '_in44': '復路(44系統)', '_out43': '往路(43系統)'},
        'takahama': {'_out': '往路(東舞鶴→高浜)', '_in': '復路(高浜→東舞鶴)'},
        'taiNohara': {'_out': '往路(東舞鶴→野原)', '_in': '復路(野原→東舞鶴)'},
        'aseku': {'_47': '大波経由(47系統)', '_48': '安岡経由(48系統)'},
        'wada': {'_out': '往路(東舞鶴→西舞鶴)', '_in': '復路(西舞鶴→東舞鶴)'},
        'joMizoshiri': {'_71': '与保呂経由(71系統)', '_72': '矢の助経由(72系統)'},
        'magura': {'_out': '往路(西舞鶴→真倉)', '_in': '復路(真倉→西舞鶴)'},
        'shinaiKita': {'_11': '東舞鶴発→雁又(11系統)', '_12': '雁又→東舞鶴(12系統)',
                       '_21': '西舞鶴発→雁又(21系統)', '_22': '雁又→西舞鶴(22系統)'},
    }
    return d.get(route_key, {}).get(suffix, suffix or '（単方向）')


def collect_directions(timetable):
    dirs = {}
    for stop_data in timetable.values():
        for entry_key in stop_data.get('routes', {}):
            parts = entry_key.split('_', 1)
            rk = parts[0]
            sf = '_' + parts[1] if len(parts) > 1 else ''
            dirs.setdefault(rk, set()).add(sf)
    return dirs


def build_excel(bus_data, timetable):
    wb = Workbook()
    wb.remove(wb.active)

    directions = collect_directions(timetable)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    for route_key, route_info in bus_data.items():
        name_ja = route_info.get('name_ja', route_key)
        color = route_info.get('color', '#333333')
        stops = route_info.get('stops', [])
        stop_names = [s['name'] for s in stops]

        route_dirs = sorted(directions.get(route_key, {''}))
        if not route_dirs:
            route_dirs = ['']

        for suffix in route_dirs:
            entry_key = f"{route_key}{suffix}"
            dir_name = get_direction_info(route_key, suffix)

            # この方向キーのデータを持つ全停留所を収集（路線順 + 路線外）
            all_stops_with_data = []
            route_stop_set = set(stop_names)

            # 1) 路線の停留所（順番通り）
            for sn in stop_names:
                tt = timetable.get(sn, {}).get('routes', {}).get(entry_key, {})
                all_stops_with_data.append({
                    'name': sn, 'in_route': True,
                    'dest': tt.get('dest', ''), 'times': tt.get('times', [])
                })

            # 2) 路線外の停留所（この方向のデータあり）
            for sn, sd in sorted(timetable.items()):
                if sn not in route_stop_set and entry_key in sd.get('routes', {}):
                    tt = sd['routes'][entry_key]
                    all_stops_with_data.append({
                        'name': sn, 'in_route': False,
                        'dest': tt.get('dest', ''), 'times': tt.get('times', [])
                    })

            # シート名
            sheet_name = f"{name_ja}_{dir_name}"[:31]
            if sheet_name in wb.sheetnames:
                sheet_name = sheet_name[:28] + '_' + suffix[-2:]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = color.lstrip('#')

            hfill = PatternFill(start_color=lighten(color, 0.6), end_color=lighten(color, 0.6), fill_type='solid')

            # Row 1: 路線情報
            ws.merge_cells('A1:B1')
            ws['A1'] = f"路線: {name_ja}"
            ws['A1'].font = Font(bold=True, size=14, color=color.lstrip('#'))
            ws['C1'] = f"方向: {dir_name}"
            ws['C1'].font = Font(bold=True, size=12)
            ws['E1'] = f"キー: {entry_key}"
            ws['E1'].font = Font(italic=True, size=10, color='888888')

            # Row 3: ヘッダー
            headers = ['No.', '停留所名', '行先表示 (dest)', '発車時刻（カンマ区切り）']
            widths = [5, 22, 28, 80]
            for ci, (h, w) in enumerate(zip(headers, widths), 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = Font(bold=True, size=10)
                c.fill = hfill
                c.border = thin
                ws.column_dimensions[chr(64+ci)].width = w

            # Row 4+: 停留所データ
            row = 4
            idx = 1
            for item in all_stops_with_data:
                ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2, value=item['name'])
                ws.cell(row=row, column=3, value=item['dest'])
                ws[f'C{row}'].font = Font(size=9, color='666666')
                ws.cell(row=row, column=4, value=', '.join(item['times']))
                ws[f'D{row}'].font = Font(name='Consolas', size=10)

                for ci in range(1, 5):
                    ws.cell(row=row, column=ci).border = thin

                # 色分け
                if not item['times']:
                    fill = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')
                    for ci in range(1, 5):
                        ws.cell(row=row, column=ci).fill = fill
                elif not item['in_route']:
                    fill = PatternFill(start_color='FFE8E8', end_color='FFE8E8', fill_type='solid')
                    for ci in range(1, 5):
                        ws.cell(row=row, column=ci).fill = fill
                    ws.cell(row=row, column=2).font = Font(color='CC0000')

                idx += 1
                row += 1

            # 凡例
            row += 1
            ws.cell(row=row, column=2, value='■ 黄色背景 = 時刻データなし（始発/終着など）')
            ws[f'B{row}'].font = Font(size=9, color='999900')
            row += 1
            ws.cell(row=row, column=2, value='■ 赤背景 = この路線の停留所リスト外（他路線経由の便）')
            ws[f'B{row}'].font = Font(size=9, color='CC0000')

    # 全停留所一覧シート
    ws = wb.create_sheet(title='全停留所一覧', index=0)
    ws.sheet_properties.tabColor = '333333'
    ws['A1'] = '全停留所一覧（路線別）'
    ws['A1'].font = Font(bold=True, size=14)
    for ci, (h, w) in enumerate(zip(['路線', 'No.', '停留所名', '緯度', '経度'], [18, 5, 22, 12, 12]), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True)
        c.border = thin
        ws.column_dimensions[chr(64+ci)].width = w

    row = 4
    for route_key, route_info in bus_data.items():
        name_ja = route_info.get('name_ja', route_key)
        color = route_info.get('color', '#333333')
        rfill = PatternFill(start_color=lighten(color, 0.7), end_color=lighten(color, 0.7), fill_type='solid')
        for idx, stop in enumerate(route_info.get('stops', []), 1):
            ws.cell(row=row, column=1, value=name_ja if idx == 1 else '')
            ws.cell(row=row, column=2, value=idx)
            ws.cell(row=row, column=3, value=stop['name'])
            ws.cell(row=row, column=4, value=stop.get('lat', ''))
            ws.cell(row=row, column=5, value=stop.get('lng', ''))
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).fill = rfill
                ws.cell(row=row, column=ci).border = thin
            row += 1
        row += 1

    return wb


def main():
    with open(HTML_PATH, 'r', encoding='utf-8') as f:
        html = f.read()
    print("index.html 読み込み中...")
    bus_data, timetable = extract_js_object(html, 'BUS_LINE_DATA'), extract_js_object(html, 'BUS_STOP_TIMETABLE')
    rc = len([k for k in bus_data if k != 'jr'])
    sc = len(timetable)
    ec = sum(len(s.get('routes', {})) for s in timetable.values())
    print(f"路線: {rc}, 停留所: {sc}, エントリ: {ec}")
    print("Excel生成中...")
    wb = build_excel(bus_data, timetable)
    wb.save(EXCEL_PATH)
    print(f"✅ 保存: {EXCEL_PATH}")
    print(f"   シート数: {len(wb.sheetnames)}")
    for n in wb.sheetnames:
        print(f"   - {n}")


if __name__ == '__main__':
    main()
